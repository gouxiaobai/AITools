from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from generate_alliance_duel_testcases import Case, HEADERS, build_cases


CORE_P1_TITLES = {
    "首次开启前1天满足个人条件玩家可见预告期入口",
    "活动进行中入口显示当日主题倒计时和每日气泡",
    "达标联盟玩家首次登录弹出预告期弹窗并支持通过入口再次打开",
    "未达标联盟玩家预告期弹窗显示条件X和红字提示",
    "无联盟玩家预告期弹窗点击加入联盟跳转联盟列表",
    "本服匹配按战力排名进入1-4、5-10、11-32号匹配池",
    "触发进行中主题任务可获得对应积分",
    "1-3档里程碑默认解锁且达标后可手动领取",
    "活动期间可查看日积分排行和总积分排行，活动结束后清空",
    "对决战况对三类满足个人条件玩家可见但无联盟/未达标玩家不可参与排行",
    "本服突袭返回我方联盟集合点使用相同落点规则",
    "联赛记录弹窗按4周分页并默认打开进行中周页签",
    "联赛前3周周日展示期仅展示联赛页签，第4周结束后入口关闭",
    "首次跨服匹配后世界聊天新增对决页签并覆盖匹配战区所有玩家",
}


TASK_TRIGGERS = [
    (1, "购买包含钻石的礼包", "触发计分、重复购买累计、跳转复用军备竞赛", "P1"),
    (2, "训练指定数量和等级士兵", "不同兵种/等级参数生效、批量训练累计", "P0"),
    (3, "使用建造加速", "分钟数累计、跨建筑使用累计", "P1"),
    (4, "使用科研加速", "分钟数累计、科技研究中/未研究中校验", "P1"),
    (5, "使用训练加速", "指定兵营与分钟数校验", "P1"),
    (6, "进行高级招募", "单次/多次招募累计", "P1"),
    (7, "单次至少消耗英雄经验", "单次阈值判定，不允许多次小额拼单", "P0"),
    (8, "消耗直升机战斗数据", "单次和累计口径确认", "P1"),
    (9, "消耗体力", "体力扣减成功才计分", "P1"),
    (10, "提升建筑战力", "建筑战力提升口径、一次升级多档累计", "P0"),
    (11, "提升科技战力", "科技完成时计分，取消/加速中不提前记分", "P0"),
    (12, "完成雷达任务", "完成态触发、前往雷达跳转正确", "P1"),
    (13, "消耗技能勋章", "消耗成功后记分、跳转正确", "P1"),
    (14, "采集粮食", "仅采集到指定资源记分，跳转到资源地正确", "P1"),
    (15, "进行军备招募", "指定招募入口触发", "P1"),
    (16, "获得UR品质的某类奖励/英雄", "品质判定准确，不混入低品质", "P1"),
    (17, "使用治疗科技/功能", "前往医院和治疗动作计分", "P1"),
    (18, "执行某类角色/部队强化", "强化成功后记分，跳转正确", "P1"),
    (19, "每消灭指定等级士兵", "PVP击杀计分口径准确", "P0"),
    (20, "采集木材", "采集目标资源正确", "P1"),
    (21, "采集铁矿", "采集目标资源正确", "P1"),
    (22, "购买直升机相关养成资源", "购买成功后计分，商城跳转正确", "P1"),
    (23, "消耗某类科技勋章", "消耗动作和计分幂等", "P1"),
    (24, "每消灭指定等级直升机单位", "目标类型和等级识别正确", "P0"),
    (25, "获得UR英雄碎片", "碎片品质判定、获得来源兼容", "P1"),
    (26, "获得SSR英雄碎片", "碎片品质判定、获得来源兼容", "P1"),
    (27, "获得SR英雄碎片", "碎片品质判定、获得来源兼容", "P1"),
    (28, "每消灭匹配对手的指定等级士兵", "只统计匹配对手，不统计其他敌对目标", "P0"),
    (29, "每损失指定等级己方士兵", "仅损失计分，不混入受伤/医院状态", "P0"),
]


DATA_PLAN = {
    "账号准备": [
        "达标联盟主测账号 16 个以上：用于本服/跨服/联赛配对、排行、联盟突袭。",
        "未达标联盟账号 4-8 个：覆盖活动可见但权限受限场景。",
        "无联盟账号 2-4 个：覆盖预告弹窗、无联盟引导、活动可见性。",
        "高主堡等级和刚达8级边界账号各至少1个：覆盖个人参与门槛边界。",
    ],
    "联盟准备": [
        "同一战区内准备排名1-4、5-10、11-32的达标联盟，验证本服三档匹配池。",
        "准备人数不足20或排名32名外联盟，用于未达标联盟补位和权限限制。",
        "准备可中途入盟、退盟、重入的测试联盟，用于联盟总积分边界。",
        "联赛场景准备至少32个联盟；如需验证完整首次联赛分层，建议准备128个联盟数据池或通过后台造数。",
    ],
    "战区与服务器准备": [
        "准备至少4个战区用于联赛开启与跨服分组验证。",
        "准备不同开服天数区间战区，用于验证k1/k2跨服分区回退逻辑。",
        "准备已配置跨服分组、漏配分组、不同区间混组3类运营配置样例。",
    ],
    "数值与奖励准备": [
        "准备可快速达成1-3档、4-6档、7-9档里程碑的积分造数方案。",
        "准备已研究/未研究67109057与67109063科技的账号，用于里程碑解锁对比。",
        "准备个人积分接近每日参与门槛、日排行最低上榜积分、联赛奖励积分门槛的边界数据。",
        "准备有/无高级迁城、有/无免费迁城、跨服冷却中3类联盟突袭账号。",
    ],
    "执行编组建议": [
        "1人负责入口/今日主题/邮件，1人负责对决战况，1人负责联盟突袭，1人负责联赛与匹配。",
        "涉及跨服和联赛的用例优先在固定时间窗集中执行，避免战区状态漂移。",
        "所有结算类用例要求记录：触发时间、服务器时间、邮件到达时间、结算截图。",
    ],
}


def numbered(lines: list[str]) -> str:
    return "\n".join(f"【{i}】{line}" for i, line in enumerate(lines, start=1))


def md_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def select_core_cases(cases: list[Case]) -> list[Case]:
    return [case for case in cases if case.level == "P0" or case.title in CORE_P1_TITLES]


def write_core_markdown(path: Path, cases: list[Case]) -> None:
    lines = [
        "# 同盟对决 P0/P1 核心回归集",
        "",
        f"- 用例数量：{len(cases)}",
        "- 选取原则：保留全部 P0 用例，并补充少量关键 P1 主流程/入口/展示期用例。",
        "",
        "| " + " | ".join(HEADERS) + " |",
        "|" + "|".join(["---"] * len(HEADERS)) + "|",
    ]
    for case in cases:
        row = [case.title, case.group, case.owner, case.precondition, numbered(case.steps), numbered(case.expected), case.level, case.case_type]
        lines.append("| " + " | ".join(md_cell(v) for v in row) + " |")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_case_workbook(path: Path, cases: list[Case], sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(HEADERS)
    bold = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = wrap
    for case in cases:
        ws.append([case.title, case.group, case.owner, case.precondition, numbered(case.steps), numbered(case.expected), case.level, case.case_type])
    for col, width in {"A": 34, "B": 34, "C": 10, "D": 32, "E": 54, "F": 54, "G": 8, "H": 10}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(path)


def write_trigger_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "触发器矩阵"
    headers = ["触发器类型", "任务描述", "专项校验重点", "优先级"]
    ws.append(headers)
    bold = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = wrap
    for row in TASK_TRIGGERS:
        ws.append(list(row))
    for col, width in {"A": 12, "B": 28, "C": 54, "D": 8}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap

    ws2 = wb.create_sheet("执行建议")
    ws2.append(["建议项", "内容"])
    ws2["A1"].font = bold
    ws2["B1"].font = bold
    suggestions = [
        ("优先级", "优先执行 2/7/10/11/19/24/28/29 这几类和战力、PVP、单次阈值相关的任务。"),
        ("幂等", "所有涉及资源扣减和PVP结算的任务类型，都应至少补1条重复请求或连续点击场景。"),
        ("跳转", "凡文档中明确给出前往入口的任务类型，都要验证跳转落点和返回后任务继续追踪。"),
        ("边界", "对“单次至少”“每消灭”“每损失”类任务，必须做刚好达标、略低于阈值、跨多次累计三种边界。"),
    ]
    for item in suggestions:
        ws2.append(list(item))
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 100
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = wrap
    wb.save(path)


def write_data_plan(path: Path) -> None:
    lines = ["# 同盟对决测试数据准备方案", ""]
    for section, items in DATA_PLAN.items():
        lines.append(f"## {section}")
        lines.append("")
        lines.extend(f"- {item}" for item in items)
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--outdir", type=Path, default=Path("generated") / "alliance_duel")
    args = parser.parse_args()
    outdir = args.outdir
    outdir.mkdir(parents=True, exist_ok=True)

    cases = build_cases()
    core_cases = select_core_cases(cases)

    write_case_workbook(outdir / "同盟对决_P0P1核心回归集.xlsx", core_cases, "核心回归集")
    write_core_markdown(outdir / "同盟对决_P0P1核心回归集.md", core_cases)
    write_trigger_workbook(outdir / "同盟对决_任务触发器专项.xlsx")
    write_data_plan(outdir / "同盟对决_测试数据准备方案.md")
    print(f"Generated core={len(core_cases)} supplement files in {outdir.resolve()}")


if __name__ == "__main__":
    main()
