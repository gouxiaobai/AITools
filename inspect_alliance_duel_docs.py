from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(r"E:\BM\plan\运营文档\活动开发\同盟对决")


def read_docx_text(path: Path) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    paragraphs: list[str] = []
    for para in root.findall(".//w:p", ns):
        texts = [node.text for node in para.findall(".//w:t", ns) if node.text]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    return "\n".join(paragraphs)


def preview_workbook(path: Path, max_rows: int = 8) -> str:
    wb = load_workbook(path, data_only=True)
    lines = [f"FILE: {path.name}"]
    for ws in wb.worksheets[:6]:
        lines.append(f"SHEET: {ws.title} rows={ws.max_row} cols={ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True):
            lines.append(repr(row))
        lines.append("---")
    return "\n".join(lines)


def main() -> None:
    docx_path = ROOT / "同盟对决.docx"
    docx_text = read_docx_text(docx_path)
    dump_path = Path("alliance_duel_docx.txt")
    dump_path.write_text(docx_text, encoding="utf-8")
    print("=== DOCX ===")
    print(docx_text)
    print(f"=== DOCX TXT WRITTEN: {dump_path.resolve()} ===")
    print("=== XLSX PREVIEW ===")
    for name in ["同盟对决-文本.xlsx", "同盟对决运营表.xlsx", "同盟对决配置表-旧版.xlsx"]:
        print(preview_workbook(ROOT / name))
        print("===")


if __name__ == "__main__":
    main()
