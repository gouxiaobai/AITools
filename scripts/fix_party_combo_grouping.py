#!/usr/bin/env python3
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook


REPLACEMENTS = {
    "派对射击|": "派对组合活动|派对射击|",
    "派对拼图|": "派对组合活动|派对拼图|",
    "派对BP|": "派对组合活动|派对BP|",
    r"派对射击\|": r"派对组合活动\|派对射击\|",
    r"派对拼图\|": r"派对组合活动\|派对拼图\|",
    r"派对BP\|": r"派对组合活动\|派对BP\|",
}


def fix_text(text: str) -> str:
    updated = text
    for old, new in REPLACEMENTS.items():
        updated = updated.replace(old, new)
    # Guard against accidental double-prefixing.
    updated = updated.replace("派对组合活动|派对组合活动|", "派对组合活动|")
    updated = updated.replace(r"派对组合活动\|派对组合活动\|", r"派对组合活动\|")
    return updated


def fix_markdown(path: Path) -> int:
    original = path.read_text(encoding="utf-8")
    updated = fix_text(original)
    if updated != original:
        path.write_text(updated, encoding="utf-8")
    return original.count("派对射击") + original.count("派对拼图") + original.count("派对BP")


def fix_workbook(path: Path) -> int:
    wb = load_workbook(path)
    changed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    updated = fix_text(cell.value)
                    if updated != cell.value:
                        cell.value = updated
                        changed += 1
    if changed:
        wb.save(path)
    return changed


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python fix_party_combo_grouping.py <markdown_path> <xlsx_path>")
        return 1

    md_path = Path(sys.argv[1])
    xlsx_path = Path(sys.argv[2])

    if not md_path.exists():
        print(f"ERROR: markdown not found: {md_path}")
        return 1
    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found: {xlsx_path}")
        return 1

    md_count = fix_markdown(md_path)
    xlsx_count = fix_workbook(xlsx_path)
    print(f"Updated markdown: {md_path}")
    print(f"Scanned markdown markers: {md_count}")
    print(f"Updated workbook cells: {xlsx_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
