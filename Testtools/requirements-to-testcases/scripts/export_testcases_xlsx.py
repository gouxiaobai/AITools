#!/usr/bin/env python3
import argparse
import json
from pathlib import Path


def safe_text(value):
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(str(x) for x in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def autosize(ws, max_width=80):
    from openpyxl.utils import get_column_letter

    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            text = safe_text(cell.value)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _normalize_type(case_type: str) -> str:
    mapping = {
        "happy": "功能",
        "negative": "异常",
        "boundary": "边界",
        "nfr": "非功能",
    }
    return mapping.get(str(case_type).strip().lower(), str(case_type))


def _steps_text(steps):
    if not isinstance(steps, list):
        return safe_text(steps)
    return "\n".join([f"【{i + 1}】{str(s)}" for i, s in enumerate(steps)])


def _find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = [safe_text(ws.cell(r, c).value).strip() for c in range(1, 20)]
        if "标题*" in row_vals and "预期结果" in row_vals:
            return r
    return 2


def _header_index(ws, header_row):
    idx = {}
    for c in range(1, ws.max_column + 1):
        name = safe_text(ws.cell(header_row, c).value).strip()
        if name:
            idx[name] = c
    return idx


def build_import_template_workbook(data, template_path: Path, group: str, owner: str, case_type_default: str):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install with: pip install openpyxl") from exc

    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    header_row = _find_header_row(ws)
    idx = _header_index(ws, header_row)

    required_cols = ["标题*", "所属分组", "维护人", "前置条件", "步骤描述", "预期结果", "用例等级", "用例类型"]
    missing = [k for k in required_cols if k not in idx]
    if missing:
        raise RuntimeError(f"Template missing required columns: {missing}")

    # 清空历史数据行，保留说明和表头。
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    row = header_row + 1
    for tc in data.get("test_cases", []):
        title = tc.get("标题", tc.get("title", ""))
        priority = tc.get("优先级", tc.get("priority", "P1"))
        preconditions = safe_text(tc.get("前置条件", tc.get("preconditions", [])))
        steps = _steps_text(tc.get("步骤", tc.get("steps", [])))
        expected = tc.get("预期结果", tc.get("expected_result", ""))
        ctype = tc.get("类型", tc.get("type", case_type_default))
        ctype = _normalize_type(ctype) if not case_type_default else case_type_default

        ws.cell(row, idx["标题*"], title)
        ws.cell(row, idx["所属分组"], group)
        ws.cell(row, idx["维护人"], owner)
        ws.cell(row, idx["前置条件"], preconditions)
        ws.cell(row, idx["步骤描述"], steps)
        ws.cell(row, idx["预期结果"], expected)
        ws.cell(row, idx["用例等级"], priority)
        ws.cell(row, idx["用例类型"], ctype)
        row += 1

    for r in ws.iter_rows(min_row=header_row + 1):
        for cell in r:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws)
    return wb


def build_workbook(data):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except Exception as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install with: pip install openpyxl") from exc

    wb = Workbook()

    # Sheet 1: test cases
    ws_tc = wb.active
    ws_tc.title = "测试用例"
    headers_tc = [
        "测试用例ID",
        "标题",
        "优先级",
        "类型",
        "前置条件",
        "步骤",
        "预期结果",
        "关联需求ID列表",
    ]
    ws_tc.append(headers_tc)

    for tc in data.get("test_cases", []):
        ws_tc.append([
            tc.get("测试用例ID", tc.get("test_case_id", "")),
            tc.get("标题", tc.get("title", "")),
            tc.get("优先级", tc.get("priority", "")),
            tc.get("类型", tc.get("type", "")),
            safe_text(tc.get("前置条件", tc.get("preconditions", []))),
            safe_text(tc.get("步骤", tc.get("steps", []))),
            tc.get("预期结果", tc.get("expected_result", "")),
            safe_text(tc.get("关联需求ID列表", tc.get("requirement_ids", []))),
        ])

    # Sheet 2: requirements
    ws_req = wb.create_sheet("requirements")
    headers_req = ["requirement_id", "source_text", "category", "source_file"]
    ws_req.append(headers_req)

    for req in data.get("requirements", []):
        ws_req.append([
            req.get("requirement_id", ""),
            req.get("source_text", ""),
            req.get("category", ""),
            req.get("source_file", ""),
        ])

    # Sheet 3: traceability
    ws_trace = wb.create_sheet("traceability")
    headers_trace = ["requirement_id", "test_case_ids"]
    ws_trace.append(headers_trace)

    for t in data.get("traceability", []):
        ws_trace.append([
            t.get("requirement_id", ""),
            safe_text(t.get("test_case_ids", [])),
        ])

    # Sheet 4: notes
    ws_notes = wb.create_sheet("notes")
    ws_notes.append(["kind", "content"])
    for item in data.get("assumptions", []):
        ws_notes.append(["assumption", safe_text(item)])
    for item in data.get("open_questions", []):
        ws_notes.append(["open_question", safe_text(item)])

    # Style
    for ws in (ws_tc, ws_req, ws_trace, ws_notes):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize(ws)

    return wb


def main():
    ap = argparse.ArgumentParser(description="Export generated testcases JSON to .xlsx")
    ap.add_argument("input", help="Input JSON (from gen_testcases.py)")
    ap.add_argument("-o", "--output", default="testcases.xlsx", help="Output .xlsx path")
    ap.add_argument(
        "--template",
        default="",
        help="Optional import template xlsx path. If provided, export fields follow that template.",
    )
    ap.add_argument("--group", default="未分组", help="所属分组")
    ap.add_argument("--owner", default="", help="维护人")
    ap.add_argument("--case-type", default="", help="用例类型（留空则自动映射）")
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)

    data = json.loads(in_path.read_text(encoding="utf-8"))
    if args.template:
        wb = build_import_template_workbook(
            data=data,
            template_path=Path(args.template),
            group=args.group,
            owner=args.owner,
            case_type_default=args.case_type,
        )
    else:
        wb = build_workbook(data)
    wb.save(out_path)

    print(f"Exported xlsx -> {out_path}")


if __name__ == "__main__":
    main()
