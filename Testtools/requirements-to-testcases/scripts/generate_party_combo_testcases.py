#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


OUTPUT_DIR = Path(r"D:\Testwork\AITools\Testtools\requirements-to-testcases\outputs")
MD_PATH = OUTPUT_DIR / "party_combo_activity_testcases.md"
XLSX_PATH = OUTPUT_DIR / "party_combo_activity_testcases.xlsx"
SOURCE_PATH = Path(r"E:\BM\plan\运营文档\活动开发\派对组合活动\派对活动.md")


SUMMARY = {
    "功能目标": "在同一主活动入口下组合打靶、拼图、BP 三个子活动，形成活动导流、消耗、奖励和排行闭环。",
    "参与入口": "主界面一级入口 id7，进入后默认打开“打靶”页签；页签固定排序为 打靶 -> 拼图 -> BP。",
    "核心规则": "打靶支持自选大奖、手动/自动射击、奖励不放回、保底与排行榜；拼图支持单块奖励、行列宝箱和大宝箱；BP 在通用战令基础上增加任务奖励道具与邮件补发。",
    "数值或次数限制": "打靶轮次上限 200；打靶基础保底示例为 5/6 次；BP 免费可领取任务奖励次数 1，付费可领取次数 2；BP 无限宝箱需求积分 50。",
    "状态变化": "活动入口含 NEW/倒计时/红点；打靶包含预告页、主页状态1/2/3、奖励记录、进度奖励、排行榜；拼图包含未完成、已完成未领取、已完成已领取；宝箱包含常态、可领奖、已领奖。",
    "奖励或消耗": "打靶消耗活动道具 33685769，奖励先进入奖励记录后手动领取；拼图和 BP 可作为打靶道具来源；排行、未领取奖励与进度奖励均涉及邮件补发。",
    "时间规则": "活动按配置控制开服天数、持续天数和结束倒计时；活动结束后触发排行发奖与补发。",
}

MODULES = [
    "活动主入口与页签",
    "打靶预告页与大奖选择",
    "打靶抽奖/自动射击/奖励记录/进度奖励/排行榜",
    "拼图任务、单块奖励、行列宝箱、大宝箱",
    "BP 任务奖励、等级奖励、无限宝箱",
    "礼包、红点、邮件补发、配置边界",
]

CONFIRMED = [
    "主活动入口显示倒计时、NEW 标签和数字红点。",
    "活动页包含 3 个固定顺序子活动，默认进入打靶页。",
    "打靶为 1 个自选大奖 + 11 个普通奖励，奖励不放回，存在大奖保底。",
    "玩家在命中大奖前可切换大奖，切换需点击“确定”才生效。",
    "拼图单块奖励、行/列宝箱奖励、大宝箱奖励都只可领取 1 次。",
    "BP 每日任务和目标任务增加任务奖励道具，不再只有积分。",
]

OPEN_QUESTIONS = [
    "打靶“完成 1 轮进度值 +1”与“每次抽中大奖，进度层数 +1”是否等价，需确认服务端最终口径。",
    "自动射击消耗“本轮抽中大奖需要的道具数量”的计算规则是否完全按服务端预演结果确定。",
    "打靶道具过期处理写为“走通用逻辑”，未明确活动结束时机、邮件兜底和展示口径。",
    "拼图大宝箱处同时出现“完成所有拼图可领取”与“点击大宝箱无交互”，需确认未完成/已完成两个状态下的真实交互。",
    "BP 购买等级、付费解锁、无限宝箱是否 100% 复用原英雄 BP 的已有规则，需要确认继承点。",
]

RISKS = [
    "打靶抽奖和奖励记录都涉及扣道具/发奖励/手动领取，最容易出现重复扣除、重复发奖或状态不同步。",
    "排行榜以轮次 + 达成时间排序，容易在同分并发结算、刷新时机和邮件发奖上出错。",
    "拼图的单块消失、行列宝箱解锁、过渡动画、红点更新存在链式状态切换风险。",
    "BP 任务奖励新增后，要同步验证免费/付费领取次数限制和通用红点、补发逻辑没有回归。",
]

AUTOMATION = [
    "接口自动化：打靶手动/自动射击的扣道具、发奖励、保底命中、奖励记录领取幂等。",
    "状态自动化：拼图块完成领奖 -> 行列宝箱解锁 -> 大宝箱解锁的状态迁移链路。",
    "定时任务自动化：活动结束后的打靶补发、拼图补发、BP 补发、排行发奖。",
    "回归自动化：主入口红点、页签红点、排行榜刷新、BP 任务奖励展示和通用道具获得弹窗。",
]


TEST_CASES = [
    {
        "id": "TC-PARTY-001",
        "module": "活动主入口",
        "sub_function": "入口展示",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "主界面新增一级入口 id7，显示倒计时、NEW 标签；有奖励可领取时显示数字红点。",
        "precondition": "活动已开启，账号满足活动开放条件。",
        "steps": "1. 进入主界面 2. 观察活动入口 3. 制造一个可领取奖励后再次观察入口",
        "expected": "入口可见；展示倒计时与 NEW 标签；存在可领奖时显示数字红点，无可领奖时红点消失。",
        "notes": "",
    },
    {
        "id": "TC-PARTY-002",
        "module": "活动主入口",
        "sub_function": "默认页签",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "活动页面包含 3 个子活动，固定排序为打靶、拼图、BP；每次点击入口默认打开打靶活动。",
        "precondition": "活动已开启。",
        "steps": "1. 点击活动入口 2. 观察页签顺序和默认落点 3. 切到其他页签后退出再重进",
        "expected": "页签顺序固定为打靶/拼图/BP；每次从主入口进入时默认落在打靶页。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-001",
        "module": "打靶",
        "sub_function": "预告页首次展示",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "活动期间默认先展示预告页；点击开始按钮后，后续活动期间不再展示预告页。",
        "precondition": "活动已开启，角色首次进入该活动。",
        "steps": "1. 首次进入打靶活动 2. 观察预告页 3. 点击开始按钮 4. 退出并再次进入活动",
        "expected": "首次进入显示预告页；点击开始后进入主页状态1；后续重进不再出现预告页。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-002",
        "module": "打靶",
        "sub_function": "预告页引导",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "玩家停留在预告页满 3 秒时，开始按钮添加特效并出现手型指引。",
        "precondition": "进入打靶预告页。",
        "steps": "1. 在预告页停留 2 秒观察 2. 继续停留至第 3 秒后观察",
        "expected": "3 秒前不出现引导；到达 3 秒时按钮出现特效和手型指引，指向按钮中心。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-003",
        "module": "打靶",
        "sub_function": "大奖选择确认",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "玩家选中新大奖后，必须点击确定按钮才会保存选中的大奖。",
        "precondition": "玩家位于打靶主页状态1。",
        "steps": "1. 点击大奖加号打开大奖弹窗 2. 选择一个大奖但不点确定直接关闭 3. 再次打开弹窗确认当前选择 4. 重新选择并点击确定",
        "expected": "未点击确定时不保存新选择；点击确定后主页展示最新大奖，并进入主页状态2。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-004",
        "module": "打靶",
        "sub_function": "未命中大奖前切换大奖",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "在未获得大奖前，玩家可随时更换大奖。",
        "precondition": "玩家已选择大奖但当前轮未命中大奖。",
        "steps": "1. 在主页状态2或状态3点击大奖图标 2. 更换为另一个大奖并确认 3. 继续射击",
        "expected": "命中大奖前允许切换大奖；后续本轮大奖判定以最新确认的大奖为准。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-005",
        "module": "打靶",
        "sub_function": "射击前引导",
        "test_type": "Boundary",
        "priority": "P2",
        "requirement": "主页未选择大奖持续 3 秒时出现手型指引，指向加号。",
        "precondition": "玩家进入主页状态1且未选择大奖。",
        "steps": "1. 停留 2 秒观察 2. 继续停留到第 3 秒观察",
        "expected": "3 秒前无指引；达到 3 秒后出现手型指引并指向大奖加号。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-006",
        "module": "打靶",
        "sub_function": "手动射击扣道具与随机奖励",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "玩家消耗活动道具随机 roll 1 个奖励；奖励不放回；奖励飞到奖励记录入口。",
        "precondition": "玩家已选择大奖，拥有至少 1 个打靶道具。",
        "steps": "1. 点击开始射击进入主页状态3 2. 执行一次手动射击 3. 观察道具数、奖励池、奖励记录入口",
        "expected": "消耗 1 个活动道具；命中 1 个奖励；已命中奖励不再留在本轮奖池；奖励飞入奖励记录入口。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-007",
        "module": "打靶",
        "sub_function": "道具不足时手动射击",
        "test_type": "Negative",
        "priority": "P0",
        "requirement": "道具不足时弹出道具获得弹窗。",
        "precondition": "玩家已进入主页状态3，活动道具为 0。",
        "steps": "1. 点击射击按钮",
        "expected": "不触发射击；不扣任何资源；弹出道具获得弹窗。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-008",
        "module": "打靶",
        "sub_function": "大奖保底",
        "test_type": "Boundary",
        "priority": "P0",
        "requirement": "达到保底次数时必中大奖；示例配置保底为第 5/6 次。",
        "precondition": "配置已生效，当前轮还未命中大奖。",
        "steps": "1. 连续射击直到保底前 1 次 2. 记录未命中大奖 3. 进行保底次数那一次射击",
        "expected": "保底前未提前命中时，达到保底次数的那次必中大奖；轮次立即结束。",
        "notes": "需结合实际 roundReward 配置逐轮验证。",
    },
    {
        "id": "TC-SHOOT-009",
        "module": "打靶",
        "sub_function": "命中大奖后的状态切换",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "命中大奖后出现获得大奖弹窗；关闭后切回主页状态2，并默认展示玩家最新选择的大奖。",
        "precondition": "当前轮即将命中大奖。",
        "steps": "1. 射击命中大奖 2. 观察大奖弹窗 3. 关闭弹窗 4. 观察回退页面",
        "expected": "出现获得大奖弹窗；关闭后回到主页状态2；显示最新大奖；存在过场动画。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-010",
        "module": "打靶",
        "sub_function": "轮次上限",
        "test_type": "Boundary",
        "priority": "P0",
        "requirement": "活动设置有最大层数 200；达到最大层数后开始射击按钮置灰，显示“已达最大层数”。",
        "precondition": "玩家当前进度达到 199 层。",
        "steps": "1. 完成当前轮命中大奖 2. 返回主页状态2 3. 观察开始射击按钮与文案",
        "expected": "达到第 200 层后按钮置灰且无交互；出现“已达最大层数”提示；不能开启新一轮。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-011",
        "module": "打靶",
        "sub_function": "自动射击完整流程",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "自动射击一次性消耗本轮抽中大奖需要的道具数量；每轮需重新点击自动射击按钮才能触发新一轮。",
        "precondition": "玩家拥有足量活动道具，当前轮未命中大奖。",
        "steps": "1. 点击自动射击按钮 2. 在确认弹窗中确认 3. 观察本轮消耗数量和结算 4. 轮次结束后再次观察按钮状态",
        "expected": "弹出二次确认；确认后一次性完成本轮直到命中大奖；道具消耗量等于本轮实际需求；下一轮需再次手动触发自动射击。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-012",
        "module": "打靶",
        "sub_function": "自动射击道具不足",
        "test_type": "Negative",
        "priority": "P0",
        "requirement": "剩余活动道具数量小于本轮需要数量时，消耗掉所有道具后弹出道具获得弹窗。",
        "precondition": "玩家剩余道具少于本轮自动命中大奖所需数量，但大于 0。",
        "steps": "1. 点击自动射击并确认 2. 观察消耗、奖励结算和弹窗",
        "expected": "已有道具被全部消耗；按已消耗次数正常结算奖励；未命中大奖则弹出道具获得弹窗。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-013",
        "module": "打靶",
        "sub_function": "跳过动画",
        "test_type": "Positive",
        "priority": "P1",
        "requirement": "勾选跳过后仅跳过手动和自动射击过程中的打靶动画；状态间过场动画不跳过。",
        "precondition": "玩家位于主页状态3。",
        "steps": "1. 勾选跳过 2. 分别执行手动射击和自动射击 3. 观察打靶动画与状态切换动画",
        "expected": "打靶过程动画被跳过，只展示奖励结果；状态切换过场动画仍正常播放。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-014",
        "module": "打靶",
        "sub_function": "奖励记录手动领取",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "奖励记录储存玩家获得的所有奖励；玩家需在该弹窗内手动点击领取奖励才会进背包。",
        "precondition": "玩家已通过打靶获得至少 1 个未领取奖励。",
        "steps": "1. 打开奖励记录弹窗 2. 点击领取 3. 检查背包和奖励记录上下区",
        "expected": "未领取奖励发入背包；奖励从未领取区移至已领取区；无重复奖励。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-015",
        "module": "打靶",
        "sub_function": "奖励记录幂等",
        "test_type": "Robustness",
        "priority": "P0",
        "requirement": "奖励记录中的奖励需要手动领取，且应避免重复领取。",
        "precondition": "奖励记录中存在 1 个可领取奖励，网络有轻微延迟。",
        "steps": "1. 连续快速点击同一个领取按钮 2. 或在请求返回前重复打开/关闭奖励记录 3. 检查背包数量和记录状态",
        "expected": "奖励只发放一次；记录最终只变为已领取一次；无重复入包。",
        "notes": "基于高风险幂等场景补充。",
    },
    {
        "id": "TC-SHOOT-016",
        "module": "打靶",
        "sub_function": "进度奖励状态与滚动",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "进度奖励有未达成/可领取/已领取三种状态；进度条默认展示 5 个半奖励并左对齐展示最新达成节点。",
        "precondition": "玩家进度接近一个奖励门槛。",
        "steps": "1. 在门槛前查看奖励状态 2. 命中大奖使进度刚好达标 3. 观察滚动位置与按钮状态 4. 领取奖励后再次观察",
        "expected": "达标前不可领；达标后变为可领取并滚动到最新节点；领取后变为已领取；已达成节点展示正确。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-017",
        "module": "打靶",
        "sub_function": "排行榜排序",
        "test_type": "Boundary",
        "priority": "P0",
        "requirement": "本服排行依据轮次高低排名；轮次相同时依据达成时间先后排名。",
        "precondition": "准备两个以上同服玩家账号。",
        "steps": "1. 让 A、B 达成不同轮次 2. 验证高轮次排前 3. 让 B、C 达成相同轮次但完成时间不同 4. 刷新排行榜",
        "expected": "轮次高者排前；轮次相同时时间更早者排前；排序稳定不抖动。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-018",
        "module": "打靶",
        "sub_function": "排行榜刷新与展示",
        "test_type": "Positive",
        "priority": "P1",
        "requirement": "点进活动界面刷新排名；未进入配置名次时显示未上榜文本；上榜时展示可领取排名奖励道具。",
        "precondition": "准备一个上榜账号和一个未上榜账号。",
        "steps": "1. 分别进入排行榜弹窗 2. 观察未上榜和上榜展示 3. 退出活动后重新进入验证刷新",
        "expected": "进入活动时刷新排名；未上榜显示未上榜文本；上榜显示对应排名奖励；列表支持上下滑动，个人排名固定。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-019",
        "module": "打靶",
        "sub_function": "数字红点",
        "test_type": "Positive",
        "priority": "P1",
        "requirement": "活动内有可领取奖励时，主界面入口、活动页签、奖励记录入口显示数字红点；领取后消失。",
        "precondition": "分别制造进度奖励、奖励记录奖励等可领取状态。",
        "steps": "1. 观察 3 处红点 2. 全部领取后再次观察",
        "expected": "有可领奖时三处显示数字红点；全部领取后红点同步消失。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-020",
        "module": "打靶",
        "sub_function": "邮件补发",
        "test_type": "Robustness",
        "priority": "P0",
        "requirement": "未手动领取的打靶奖励、未手动领取的进度奖励、排行奖励均在活动结束后通过邮件下发。",
        "precondition": "活动结束前保留一份未领取打靶奖励、一份未领取进度奖励，并让账号处于某个排行奖励档位。",
        "steps": "1. 不领取奖励等待活动结束 2. 检查邮件和背包 3. 重新登录再次检查",
        "expected": "三类奖励在活动结束后通过邮件准确补发；邮件内容与档位一致；不重复发放。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-021",
        "module": "打靶",
        "sub_function": "礼包入口与礼包限制",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "礼包界面复用幸运转盘模板；存在每日免费礼包与多档礼包顺序/每日最大购买次数配置。",
        "precondition": "活动开启，礼包配置已生效。",
        "steps": "1. 打开礼包界面 2. 验证免费礼包可领取 3. 验证同类礼包顺序和多层购买顺序 4. 达到每日购买上限后再次尝试购买",
        "expected": "礼包页正常展示；免费礼包可领取且仅按配置次数领取；多层礼包购买顺序正确；到达每日上限后不能继续购买。",
        "notes": "",
    },
    {
        "id": "TC-SHOOT-022",
        "module": "打靶",
        "sub_function": "道具来源弹窗",
        "test_type": "Positive",
        "priority": "P2",
        "requirement": "道具获得弹窗为通用道具获取来源弹窗，包含拼图和 BP 两个渠道。",
        "precondition": "打靶道具不足，触发道具获得弹窗。",
        "steps": "1. 打开道具获得弹窗 2. 观察来源入口和跳转",
        "expected": "弹窗展示拼图和 BP 两个来源入口；点击后跳转到对应模块。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-001",
        "module": "拼图",
        "sub_function": "单块奖励一次性领取",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "玩家达成拼图任务要求可领取单个拼图奖励，每个拼图奖励只可领取 1 次。",
        "precondition": "某个拼图任务已达成但未领奖。",
        "steps": "1. 点击该拼图 2. 在弹窗中点击领取 3. 再次点击原位置",
        "expected": "奖励成功领取一次；拼图领奖后消失；不能再次领取同一拼图奖励。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-002",
        "module": "拼图",
        "sub_function": "未完成任务且有跳转",
        "test_type": "Positive",
        "priority": "P1",
        "requirement": "未完成拼图若配置了跳转路径，弹窗显示前往按钮，玩家可跳转到指定任务页面。",
        "precondition": "存在一个未完成且配置了跳转路径的拼图任务。",
        "steps": "1. 点击未完成拼图 2. 观察弹窗按钮 3. 点击前往",
        "expected": "弹窗显示前往按钮；点击后跳转到配置的任务页面。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-003",
        "module": "拼图",
        "sub_function": "未完成任务且无跳转",
        "test_type": "Negative",
        "priority": "P1",
        "requirement": "未完成拼图若未配置跳转路径，弹窗显示置灰领取按钮，点击无任何交互反馈。",
        "precondition": "存在一个未完成且未配置跳转路径的拼图任务。",
        "steps": "1. 点击未完成拼图 2. 观察按钮状态 3. 点击置灰按钮",
        "expected": "显示置灰领取按钮；点击后无跳转、无领奖、无异常提示。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-004",
        "module": "拼图",
        "sub_function": "任务进度文本",
        "test_type": "Boundary",
        "priority": "P2",
        "requirement": "任务描述文本里展示玩家当前数值与任务数值，例如累计登录 5 次（1/5）。",
        "precondition": "存在一个可累计进度的拼图任务。",
        "steps": "1. 在 0 进度时查看弹窗 2. 累积到中间进度再查看 3. 达成目标后再次查看",
        "expected": "弹窗持续展示当前值/目标值；中间进度变化准确；达成后数值与状态同步。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-005",
        "module": "拼图",
        "sub_function": "行列宝箱解锁",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "任意 1 行或 1 列的所有拼图块均完成领奖并消失后，对应宝箱立即变为可领奖状态。",
        "precondition": "某一行或某一列仅剩最后 1 个拼图未领奖。",
        "steps": "1. 领取最后一个拼图奖励 2. 观察行/列宝箱状态变化",
        "expected": "最后一个拼图消失后，目标行/列宝箱立即变为可领奖状态。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-006",
        "module": "拼图",
        "sub_function": "行列宝箱过渡动画",
        "test_type": "Robustness",
        "priority": "P1",
        "requirement": "在最后 1 个拼图消失与宝箱变可领奖之间增加过渡动画做平滑衔接和引导。",
        "precondition": "某个行/列宝箱即将被解锁。",
        "steps": "1. 领取最后一个相关拼图奖励 2. 观察拼图消失和宝箱点亮过程",
        "expected": "拼图消失到宝箱可领之间存在过渡动画；动画完成后宝箱进入可领奖状态；无卡死或状态错位。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-007",
        "module": "拼图",
        "sub_function": "小宝箱三态",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "小宝箱包含常态、可领奖、已领奖三种状态；常态点击打开气泡，可领奖带红点和动效，已领奖无交互。",
        "precondition": "分别准备未达成、可领奖、已领奖三种小宝箱。",
        "steps": "1. 依次点击三种状态的小宝箱 2. 观察交互和动画",
        "expected": "常态点击打开气泡；可领奖带红点和动效且可领奖；已领奖为打开状态且无交互。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-008",
        "module": "拼图",
        "sub_function": "大宝箱完成链路",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "完成所有拼图后，玩家可领取大宝箱奖励；大宝箱奖励只可领取 1 次。",
        "precondition": "除最后一个拼图外，其余拼图均已完成领奖。",
        "steps": "1. 领取最后一个拼图奖励 2. 观察大宝箱状态 3. 领取大宝箱 4. 再次尝试交互",
        "expected": "全部拼图完成后大宝箱可领取；奖励只能领取一次；领取后状态正确更新。",
        "notes": "需和文档“点击大宝箱无交互”口径一起确认。",
    },
    {
        "id": "TC-PUZZLE-009",
        "module": "拼图",
        "sub_function": "红点传染",
        "test_type": "Positive",
        "priority": "P1",
        "requirement": "有奖励可领取时，主界面入口和活动页签显示数字红点；对应拼图/宝箱显示普通红点。",
        "precondition": "制造一个可领取拼图奖励和一个可领取宝箱奖励。",
        "steps": "1. 观察主界面入口、页签、局部拼图/宝箱红点 2. 领取奖励后再次观察",
        "expected": "主界面入口和页签显示数字红点；局部目标显示普通红点；全部领取后红点同步消失。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-010",
        "module": "拼图",
        "sub_function": "邮件补发",
        "test_type": "Robustness",
        "priority": "P0",
        "requirement": "活动期间未领取的拼图奖励或宝箱奖励通过邮件补发给玩家。",
        "precondition": "活动结束前保留若干未领取拼图/宝箱奖励。",
        "steps": "1. 不领取奖励直至活动结束 2. 检查邮件和背包",
        "expected": "未领取拼图奖励和宝箱奖励通过邮件补发；奖励内容与原配置一致；不重复补发。",
        "notes": "",
    },
    {
        "id": "TC-PUZZLE-011",
        "module": "拼图",
        "sub_function": "box 配置映射",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "puzzleBoxReward 通过 unlockPuzzleId 映射宝箱解锁条件。",
        "precondition": "已知某个 boxId 的 unlockPuzzleId 配置，例如 101 -> 1,2,3。",
        "steps": "1. 仅完成其中部分拼图 2. 观察宝箱状态 3. 完成全部配置拼图后再次观察",
        "expected": "未完成全部配置拼图前宝箱不可领取；完成全部配置拼图后准确解锁对应宝箱。",
        "notes": "",
    },
    {
        "id": "TC-BP-001",
        "module": "BP",
        "sub_function": "任务奖励新增道具",
        "test_type": "Positive",
        "priority": "P0",
        "requirement": "每日任务和目标任务增加任务奖励，不只给积分。",
        "precondition": "存在一个可快速完成的 BP 任务。",
        "steps": "1. 完成任务 2. 领取任务奖励 3. 检查积分和道具",
        "expected": "任务领取后同时获得积分和 taskRewards 道具，不再只有积分。",
        "notes": "",
    },
    {
        "id": "TC-BP-002",
        "module": "BP",
        "sub_function": "完成任务提示",
        "test_type": "Positive",
        "priority": "P1",
        "requirement": "完成任务的获得积分提示改为通用道具获得界面。",
        "precondition": "准备一个可完成的 BP 任务。",
        "steps": "1. 完成任务并领取 2. 观察获得提示界面",
        "expected": "弹出通用道具获得界面，正确展示获得的积分和道具。",
        "notes": "",
    },
    {
        "id": "TC-BP-003",
        "module": "BP",
        "sub_function": "免费/付费领取次数限制",
        "test_type": "Boundary",
        "priority": "P0",
        "requirement": "配置中 freeTimes=1，chargeTimes=2；taskRewards 受该配置生效。",
        "precondition": "准备免费状态和付费状态账号各 1 个，均有可领取 BP 任务奖励。",
        "steps": "1. 免费账号连续领取任务奖励至上限 2. 付费账号连续领取任务奖励至上限 3. 超限后再次尝试领取",
        "expected": "免费账号最多领取 1 次，付费账号最多领取 2 次；超限后不可继续领取积分和任务道具。",
        "notes": "",
    },
    {
        "id": "TC-BP-004",
        "module": "BP",
        "sub_function": "购买等级消耗钻石",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "购买等级消耗钻石，示例配置 gemsLvlPay=200。",
        "precondition": "账号钻石分别准备为 199 和 200。",
        "steps": "1. 以 199 钻石尝试购买等级 2. 以 200 钻石再次尝试",
        "expected": "199 钻石时购买失败且不升级；200 钻石时购买成功并正确扣钻。",
        "notes": "",
    },
    {
        "id": "TC-BP-005",
        "module": "BP",
        "sub_function": "无限宝箱",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "无限宝箱需求积分为 50，达到后可领取对应奖励。",
        "precondition": "玩家当前积分为 49。",
        "steps": "1. 完成任务获得 1 点以上积分 2. 检查无限宝箱状态并领取",
        "expected": "49 分时不可领；达到 50 分后无限宝箱可领取；奖励内容与配置一致。",
        "notes": "",
    },
    {
        "id": "TC-BP-006",
        "module": "BP",
        "sub_function": "等级奖励展示",
        "test_type": "Positive",
        "priority": "P1",
        "requirement": "bp_award 配置包含普通奖励和付费奖励两条轨道。",
        "precondition": "准备普通状态和付费解锁状态账号。",
        "steps": "1. 提升到至少 2 级 2. 观察等级奖励展示和领取结果",
        "expected": "普通账号只能领取普通轨奖励；付费账号可额外领取 chargeAward；展示内容和配置一致。",
        "notes": "",
    },
    {
        "id": "TC-BP-007",
        "module": "BP",
        "sub_function": "红点与邮件补发",
        "test_type": "Robustness",
        "priority": "P0",
        "requirement": "BP 红点走通用逻辑并传染到页签和主界面图标；活动结束后未领取奖励通过邮件补发。",
        "precondition": "制造一个未领取的 BP 奖励。",
        "steps": "1. 观察主界面入口和 BP 页签红点 2. 不领取直到活动结束 3. 检查邮件",
        "expected": "可领奖时红点传染到页签和主界面图标；活动结束后未领取 BP 奖励通过邮件补发。",
        "notes": "",
    },
    {
        "id": "TC-CROSS-001",
        "module": "跨模块",
        "sub_function": "活动开放条件",
        "test_type": "Boundary",
        "priority": "P1",
        "requirement": "打靶、拼图、BP 均通过配置控制开服活动、开始开服天数、持续时间和一级入口。",
        "precondition": "准备满足和不满足开服天数/城堡等级条件的账号。",
        "steps": "1. 在开放前观察入口 2. 到达开放条件后再次观察 3. 活动结束后再次检查",
        "expected": "未满足条件时入口不可见或不可进；满足条件后显示；活动结束后入口按配置关闭。",
        "notes": "",
    },
    {
        "id": "TC-CROSS-002",
        "module": "跨模块",
        "sub_function": "页签切换状态保持",
        "test_type": "Positive",
        "priority": "P2",
        "requirement": "三子活动在同一活动容器内切换展示。",
        "precondition": "三个子活动均已解锁。",
        "steps": "1. 在打靶页切到拼图/BP 再切回 2. 观察各自局部状态和红点刷新",
        "expected": "页签切换正常；各模块状态分别保持；红点和奖励状态刷新正确。",
        "notes": "",
    },
    {
        "id": "TC-CROSS-003",
        "module": "跨模块",
        "sub_function": "多来源导流",
        "test_type": "Positive",
        "priority": "P2",
        "requirement": "打靶道具来源包含拼图、BP 和礼包；形成活动联动。",
        "precondition": "打靶道具不足。",
        "steps": "1. 从道具来源弹窗分别进入拼图/BP/礼包 2. 完成对应行为获得打靶道具 3. 返回打靶活动",
        "expected": "三个来源跳转链路可用；获得的道具可正常回流到打靶活动使用。",
        "notes": "",
    },
    {
        "id": "TC-CROSS-004",
        "module": "跨模块",
        "sub_function": "倒计时与结束结算",
        "test_type": "Boundary",
        "priority": "P0",
        "requirement": "主入口与活动名称区域展示活动结束倒计时；活动结束后触发排行发奖和补发。",
        "precondition": "活动接近结束时间。",
        "steps": "1. 在结束前观察倒计时 2. 等待活动结束 3. 再次进入主入口和子活动 4. 检查邮件",
        "expected": "倒计时准确归零；活动结束后入口/子活动状态按规则关闭；排行与未领奖补发邮件正常触发。",
        "notes": "",
    },
]


def markdown_table(rows: list[dict]) -> str:
    headers = [
        "用例ID",
        "模块",
        "子功能",
        "需求依据",
        "前置条件",
        "测试步骤",
        "预期结果",
        "假设/备注",
    ]
    lines = [
        "| " + " | ".join(headers) + " |",
        "|--------|------|--------|----------|--------|----------|----------|----------|----------|-----------|",
    ]
    for row in rows:
        values = [
            row["id"],
            row["module"],
            row["sub_function"],
            row["requirement"],
            row["precondition"],
            row["steps"],
            row["expected"],
            row["notes"] or "无",
        ]
        safe = [str(v).replace("\n", "<br>").replace("|", "\\|") for v in values]
        lines.append("| " + " | ".join(safe) + " |")
    return "\n".join(lines)


def build_markdown() -> str:
    parts = [
        "# 派对组合活动测试用例",
        "",
        f"来源文档：`{SOURCE_PATH}`",
        "",
        "**文档解读摘要：**",
    ]
    for key, value in SUMMARY.items():
        parts.append(f"- {key}：{value}")
    parts.extend(
        [
            "",
            "**涉及模块：**",
        ]
    )
    parts.extend([f"- {item}" for item in MODULES])
    parts.extend(
        [
            "",
            "**需求确认项：**",
            f"- 已确认：{'；'.join(CONFIRMED)}",
            f"- 待确认：{'；'.join(OPEN_QUESTIONS)}",
            f"- 基于经验补充的风险：{'；'.join(RISKS)}",
            "",
            "**测试用例：**",
            "",
            markdown_table(TEST_CASES),
            "",
            "**覆盖率总结：** 覆盖了主入口、页签、打靶抽奖与排行、拼图奖励链路、BP 奖励链路、红点、礼包、邮件补发和活动结束结算。",
            f"**潜在风险点：** {'；'.join(RISKS)}",
            f"**自动化建议：** {'；'.join(AUTOMATION)}",
            f"**需要补充的信息：** {'；'.join(OPEN_QUESTIONS)}",
            "",
        ]
    )
    return "\n".join(parts)


def autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "test_cases"
    headers = [
        "用例ID",
        "模块",
        "子功能",
        "需求依据",
        "前置条件",
        "测试步骤",
        "预期结果",
        "假设/备注",
    ]
    ws.append(headers)
    for row in TEST_CASES:
        ws.append(
            [
                row["id"],
                row["module"],
                row["sub_function"],
                row["requirement"],
                row["precondition"],
                row["steps"],
                row["expected"],
                row["notes"],
            ]
        )

    summary = wb.create_sheet("summary")
    summary.append(["字段", "内容"])
    for key, value in SUMMARY.items():
        summary.append([key, value])
    summary.append(["涉及模块", "；".join(MODULES)])
    summary.append(["已确认", "；".join(CONFIRMED)])
    summary.append(["待确认", "；".join(OPEN_QUESTIONS)])
    summary.append(["风险", "；".join(RISKS)])
    summary.append(["自动化建议", "；".join(AUTOMATION)])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize(sheet)

    return wb


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    MD_PATH.write_text(build_markdown(), encoding="utf-8")
    wb = build_workbook()
    wb.save(XLSX_PATH)
    print(f"Generated: {MD_PATH}")
    print(f"Generated: {XLSX_PATH}")


if __name__ == "__main__":
    main()
