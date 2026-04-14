"""Utilities for loading automation suites from .xlsx sheets."""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.comments import Comment

REQUIRED_COLUMNS = {"case_name", "step_no", "action"}


def _normalize_key(k: str) -> str:
    return (k or "").strip().lower()


def _to_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    s = str(value).strip().lower()
    if s == "":
        return default
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _parse_json_field(raw: Any, field: str, row_no: int, default: Any) -> Any:
    import json

    text = str(raw or "").strip()
    if not text:
        return default
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        raise ValueError(f"row {row_no}: invalid {field}: {e}") from e


def _rows_from_worksheet(ws) -> list[dict[str, Any]]:
    it = ws.iter_rows(values_only=True)
    headers = next(it, None)
    if not headers:
        raise ValueError("XLSX has no header row")
    keys = [_normalize_key(str(h) if h is not None else "") for h in headers]
    if not any(keys):
        raise ValueError("XLSX header row is empty")

    rows: list[dict[str, Any]] = []
    for values in it:
        row = {keys[i]: values[i] if i < len(values) else "" for i in range(len(keys))}
        rows.append(row)
    return rows


def load_rows_from_xlsx_path(path: str | Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return _rows_from_worksheet(wb.active)
    finally:
        wb.close()


def load_rows_from_xlsx_bytes(data: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return _rows_from_worksheet(wb.active)
    finally:
        wb.close()


def build_suite(rows: list[dict[str, Any]], default_suite_name: str) -> dict[str, Any]:
    if not rows:
        raise ValueError("input has no data rows")

    missing = REQUIRED_COLUMNS - set(rows[0].keys())
    if missing:
        raise ValueError(f"missing required columns: {', '.join(sorted(missing))}")

    suite_name = default_suite_name
    cases: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    case_order: list[str] = []

    for i, row in enumerate(rows, start=2):
        if not _to_bool(row.get("enabled"), default=True):
            continue

        case_name = str(row.get("case_name") or "").strip()
        step_no_raw = str(row.get("step_no") or "").strip()
        action_raw = str(row.get("action") or "").strip()

        if not case_name and not step_no_raw and not action_raw:
            continue
        if not case_name:
            raise ValueError(f"row {i}: case_name is required")
        if not step_no_raw:
            raise ValueError(f"row {i}: step_no is required")
        if not action_raw:
            raise ValueError(f"row {i}: action is required")

        try:
            step_no = int(step_no_raw)
        except ValueError as e:
            raise ValueError(f"row {i}: step_no must be int, got {step_no_raw!r}") from e

        try:
            action = int(action_raw)
        except ValueError as e:
            raise ValueError(f"row {i}: action must be int, got {action_raw!r}") from e

        row_suite_name = str(row.get("suite_name") or "").strip()
        if row_suite_name:
            suite_name = row_suite_name

        params = _parse_json_field(row.get("params_json"), "params_json", i, {})
        if not isinstance(params, dict):
            raise ValueError(f"row {i}: params_json must be JSON object")

        expect = _parse_json_field(row.get("expect_json"), "expect_json", i, [])
        if isinstance(expect, dict):
            expect = [expect]
        if not isinstance(expect, list):
            raise ValueError(f"row {i}: expect_json must be JSON array or object")

        timeout_raw = str(row.get("timeout_ms") or "").strip()
        timeout_ms = int(timeout_raw) if timeout_raw else 5000

        resp_raw = str(row.get("response_action") or "").strip()
        response_action = int(resp_raw) if resp_raw else None

        step: dict[str, Any] = {
            "action": action,
            "params": params,
            "timeout_ms": timeout_ms,
            "expect": expect,
        }
        if response_action is not None:
            step["response_action"] = response_action

        if case_name not in cases:
            cases[case_name] = []
            case_order.append(case_name)
        cases[case_name].append((step_no, step))

    out_cases = []
    for cn in case_order:
        sorted_steps = [s for _, s in sorted(cases[cn], key=lambda x: x[0])]
        out_cases.append({"name": cn, "steps": sorted_steps})

    if not out_cases:
        raise ValueError("no enabled rows to build cases")

    return {"name": suite_name, "cases": out_cases}


def load_suite_from_xlsx_path(path: str | Path, default_suite_name: str | None = None) -> dict[str, Any]:
    p = Path(path)
    rows = load_rows_from_xlsx_path(p)
    return build_suite(rows, default_suite_name=default_suite_name or p.stem)


def load_suite_from_xlsx_bytes(data: bytes, default_suite_name: str) -> dict[str, Any]:
    rows = load_rows_from_xlsx_bytes(data)
    return build_suite(rows, default_suite_name=default_suite_name)


def build_template_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        ws.title = "suite"
        headers = [
            "suite_name",
            "case_name",
            "step_no",
            "action",
            "params_json",
            "timeout_ms",
            "response_action",
            "expect_json",
            "enabled",
            "remark",
        ]
        ws.append(headers)
        ws.append(
            [
                "sample_suite",
                "case_1",
                1,
                50006,
                "{}",
                5000,
                50006,
                '[{"type":"equals","path":"commResp.code","value":0}]',
                1,
                "sample row",
            ]
        )
        header_comments = {
            "suite_name": "套件名。可选；如填写，后续行可复用同一名称。",
            "case_name": "用例名。必填；同名行会按 step_no 聚合成一个 case。",
            "step_no": "步骤序号。必填；整数，建议从 1 开始递增。",
            "action": "请求 Action。必填；整数。",
            "params_json": "请求参数 JSON。可选；必须是 JSON 对象，例如 {} 或 {\"k\":1}。",
            "timeout_ms": "超时毫秒。可选；留空默认 5000。",
            "response_action": "期望响应 Action。可选；整数，留空表示不限制。",
            "expect_json": "断言列表。可选；JSON 数组或对象（对象会自动转单元素数组）。",
            "enabled": "是否启用本行。可选；1/true/yes/on 为启用，0/false/no/off 为禁用。",
            "remark": "备注。可选；仅用于说明，不参与执行。",
        }
        for idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.comment = Comment(header_comments[key], "GameProxyTool")

        ws_guide = wb.create_sheet("说明")
        ws_guide.append(["字段", "是否必填", "类型/格式", "填写规则", "示例"])
        guide_rows = [
            ["suite_name", "否", "string", "建议整套用例保持一致；为空时用文件名作为套件名。", "mail_suite"],
            ["case_name", "是", "string", "同名代表同一个 case。", "case_login"],
            ["step_no", "是", "int", "同一 case 内按升序执行。", "1"],
            ["action", "是", "int", "请求 action 编号。", "50006"],
            ["params_json", "否", "JSON object", "必须为 JSON 对象，不能是数组/字符串。", "{\"maxMailId\":0,\"mailSyncVersion\":0}"],
            ["timeout_ms", "否", "int", "超时时间（毫秒），为空默认 5000。", "3000"],
            ["response_action", "否", "int", "期望响应 action；为空不限制。", "50006"],
            ["expect_json", "否", "JSON array/object", "支持数组或对象；对象会转为单条断言。", "[{\"type\":\"equals\",\"path\":\"commResp.code\",\"value\":0}]"],
            ["enabled", "否", "bool", "支持 1/0、true/false、yes/no、on/off。", "1"],
            ["remark", "否", "string", "注释信息，不参与执行。", "首包校验"],
        ]
        for row in guide_rows:
            ws_guide.append(row)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    finally:
        wb.close()
