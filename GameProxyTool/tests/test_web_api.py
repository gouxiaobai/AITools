import threading
import time
import unittest
import os
import shutil
import uuid
import io

import openpyxl

from store import PacketStore
from web import create_app


class _FakeProxy:
    def __init__(self, store: PacketStore):
        self.store = store
        self._next = 1

    def active_count(self):
        return 1

    def inject(self, msg, conn_index=0):
        sid = self._next
        self._next += 1
        action = int(msg["action"])

        def _respond():
            time.sleep(0.02)
            self.store.add("S2C", b"\x02", {"commResp": {"action": action, "code": 0}})

        threading.Thread(target=_respond, daemon=True).start()
        return sid


class WebApiTests(unittest.TestCase):
    def _make_reports_dir(self):
        path = os.path.join("tests_tmp", f"reports_{uuid.uuid4().hex[:8]}")
        os.makedirs(path, exist_ok=True)
        return path

    def test_test_run_flow(self):
        store = PacketStore()
        proxy = _FakeProxy(store)
        td = self._make_reports_dir()
        try:
            app = create_app(store, proxy, reports_dir=td)
            client = app.test_client()

            suite = {
                "name": "api_suite",
                "steps": [
                    {
                        "action": 50006,
                        "timeout_ms": 500,
                        "expect": [{"type": "equals", "path": "commResp.code", "value": 0}],
                    }
                ],
            }
            resp = client.post("/api/test/run", json={"suite": suite})
            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertTrue(data["ok"])
            run_id = data["run"]["run_id"]

            # Poll until finished.
            deadline = time.time() + 3
            status = None
            while time.time() < deadline:
                r = client.get(f"/api/test/run/{run_id}")
                self.assertEqual(r.status_code, 200)
                status = r.get_json()
                if status["status"] != "running":
                    break
                time.sleep(0.05)
            self.assertIsNotNone(status)
            self.assertEqual(status["status"], "passed")

            runs = client.get("/api/test/runs").get_json()
            self.assertGreaterEqual(len(runs), 1)
        finally:
            shutil.rmtree(td, ignore_errors=True)

    def test_upload_xlsx_run_flow(self):
        store = PacketStore()
        proxy = _FakeProxy(store)
        td = self._make_reports_dir()
        try:
            app = create_app(store, proxy, reports_dir=td)
            client = app.test_client()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([
                "suite_name",
                "case_name",
                "step_no",
                "action",
                "params_json",
                "timeout_ms",
                "response_action",
                "expect_json",
                "enabled",
                "remark",
            ])
            ws.append([
                "upload_suite",
                "case1",
                1,
                50006,
                "{}",
                500,
                50006,
                '[{"type":"equals","path":"commResp.code","value":0}]',
                1,
                "ok",
            ])
            bio = io.BytesIO()
            wb.save(bio)
            wb.close()
            bio.seek(0)

            resp = client.post(
                "/api/test/upload-run",
                data={"conn": "0", "file": (bio, "upload_suite.xlsx")},
                content_type="multipart/form-data",
            )
            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertTrue(data["ok"])
            run_id = data["run"]["run_id"]

            deadline = time.time() + 3
            status = None
            while time.time() < deadline:
                r = client.get(f"/api/test/run/{run_id}")
                self.assertEqual(r.status_code, 200)
                status = r.get_json()
                if status["status"] != "running":
                    break
                time.sleep(0.05)
            self.assertIsNotNone(status)
            self.assertEqual(status["status"], "passed")
        finally:
            shutil.rmtree(td, ignore_errors=True)

    def test_download_template(self):
        store = PacketStore()
        proxy = _FakeProxy(store)
        td = self._make_reports_dir()
        try:
            app = create_app(store, proxy, reports_dir=td)
            client = app.test_client()

            resp = client.get("/api/test/template.xlsx")
            self.assertEqual(resp.status_code, 200)
            self.assertIn(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                resp.headers.get("Content-Type", ""),
            )
            self.assertIn("attachment;", resp.headers.get("Content-Disposition", ""))
            self.assertGreater(len(resp.data), 0)

            wb = openpyxl.load_workbook(io.BytesIO(resp.data), read_only=True, data_only=True)
            try:
                ws = wb.active
                headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
                self.assertIn("说明", wb.sheetnames)
                ws_guide = wb["说明"]
                guide_headers = [str(c.value) if c.value is not None else "" for c in next(ws_guide.iter_rows(max_row=1))]
            finally:
                wb.close()
            self.assertEqual(
                headers,
                [
                    "suite_name",
                    "case_name",
                    "step_no",
                    "action",
                    "params_json",
                    "timeout_ms",
                    "response_action",
                    "expect_json",
                    "enabled",
                    "remark",
                ],
            )
            self.assertEqual(guide_headers, ["字段", "是否必填", "类型/格式", "填写规则", "示例"])
        finally:
            shutil.rmtree(td, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
