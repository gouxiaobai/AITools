from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REQ_PATH = Path(r"E:\BM\plan\运营文档\活动开发\派对组合活动\派对活动.md")
OURS_PATH = Path(r"E:\BM\plan\运营文档\活动开发\派对组合活动\派对组合活动_测试用例.xlsx")
BASELINE_PATH = Path(r"D:\Testwork\测试用例\派对组合活动\party_combo_activity_testcases.xlsx")


@dataclass
class CaseRow:
    title: str
    group: str
    precondition: str
    steps: str
    expected: str
    level: str


def read_cases(path: Path) -> list[CaseRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {name: i + 1 for i, name in enumerate(headers)}
    rows: list[CaseRow] = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, idx.get("标题*", 1)).value
        if not title:
            continue
        rows.append(
            CaseRow(
                title=str(title),
                group=str(ws.cell(r, idx.get("所属分组", 2)).value or ""),
                precondition=str(ws.cell(r, idx.get("前置条件", 4)).value or ""),
                steps=str(ws.cell(r, idx.get("步骤描述", 5)).value or ""),
                expected=str(ws.cell(r, idx.get("预期结果", 6)).value or ""),
                level=str(ws.cell(r, idx.get("用例等级", 7)).value or ""),
            )
        )
    return rows


def count_by_prefix(rows: Iterable[CaseRow], prefixes: list[str]) -> dict[str, int]:
    result = {}
    for prefix in prefixes:
        result[prefix] = sum(1 for row in rows if row.group.startswith(prefix))
    return result


def count_keywords(rows: Iterable[CaseRow], keywords: dict[str, list[str]]) -> dict[str, int]:
    result: dict[str, int] = {}
    for label, words in keywords.items():
        result[label] = sum(
            1
            for row in rows
            if any(word in row.title or word in row.group or word in row.steps or word in row.expected for word in words)
        )
    return result


def top_groups(rows: Iterable[CaseRow]) -> list[tuple[str, int]]:
    counter = Counter(row.group for row in rows)
    return counter.most_common(12)


def levels(rows: Iterable[CaseRow]) -> dict[str, int]:
    counter = Counter(row.level for row in rows)
    return dict(counter)


def main() -> None:
    ours = read_cases(OURS_PATH)
    baseline = read_cases(BASELINE_PATH)
    req_text = REQ_PATH.read_text(encoding="utf-8")

    group_prefixes = [
        "派对组合活动",
        "派对射击",
        "派对拼图",
        "派对BP",
    ]
    keyword_sets = {
        "入口/页签": ["入口", "页签", "默认打开", "主界面"],
        "红点": ["红点"],
        "预告/引导": ["预告", "引导"],
        "大奖选择": ["大奖", "选择"],
        "手动射击": ["射击", "命中大奖", "保底", "奖池"],
        "自动射击": ["自动射击"],
        "进度奖励": ["进度", "层数"],
        "奖励记录": ["奖励记录", "未领取", "已领取"],
        "排行": ["排行", "排行榜", "排名"],
        "礼包/来源": ["礼包", "来源", "购买"],
        "拼图任务": ["拼图", "任务弹窗"],
        "宝箱": ["宝箱"],
        "BP": ["BP", "战令", "任务奖励", "无限宝箱"],
        "结束补发": ["补发", "邮件", "结束"],
        "并发/幂等": ["连续点击", "重复", "并发", "不重复", "仅发放一次"],
    }

    print("REQ_HAS", {
        "组合活动": "3个子活动" in req_text,
        "打靶": "## 打靶" in req_text,
        "拼图": "## 拼图" in req_text,
        "bp": "## bp" in req_text,
        "奖励记录": "奖励记录" in req_text,
        "自动射击": "自动射击" in req_text,
        "排行榜": "排行榜" in req_text,
        "邮件补发": "补发邮件" in req_text or "通过邮件补发" in req_text,
    })

    for name, rows in [("OURS", ours), ("BASELINE", baseline)]:
        print("===", name, "===")
        print("TOTAL_CASES", len(rows))
        print("LEVELS", levels(rows))
        print("GROUP_PREFIX", count_by_prefix(rows, group_prefixes))
        print("KEYWORDS", count_keywords(rows, keyword_sets))
        print("TOP_GROUPS", top_groups(rows))
        print("SAMPLE_TITLES", [row.title for row in rows[:8]])


if __name__ == "__main__":
    main()
